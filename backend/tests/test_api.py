"""End-to-end API tests: upload, async job lifecycle, exports.

TestClient executes FastAPI background tasks synchronously after each
response, so a job created via POST is already terminal on the next GET.
"""
from __future__ import annotations

import io
import sqlite3
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.core.config import get_settings
from app.db.base import Base
from app.db.models import CleaningJob, JobStatus, Report
from app.db.session import SessionLocal, engine
from app.io import writers
from app.main import app
from app.repositories.job_repository import JobRepository
from app.services.cleaning_service import CleaningService

MESSY_CSV = (
    b" Customer ID , Full Name ,e-mail\n"
    b"1,Anna,a@x.pl\n"
    b"2,Bartek,\n"
    b"2,Bartek,\n"
    b"3,,c@x.pl\n"
)


@pytest.fixture(scope="module")
def client() -> TestClient:
    Base.metadata.create_all(engine)
    with TestClient(app) as test_client:
        yield test_client


def _upload(client: TestClient, content: bytes = MESSY_CSV, name: str = "klienci.csv"):
    return client.post(
        "/datasets/upload", files={"file": (name, content, "text/csv")}
    )


def test_upload_returns_metadata(client: TestClient) -> None:
    res = _upload(client)
    assert res.status_code == 201
    body = res.json()
    assert body["row_count"] == 4
    assert body["column_count"] == 3


def test_preview_stats_insights_served_from_cache(client: TestClient) -> None:
    dataset_id = _upload(client).json()["id"]

    preview = client.get(f"/datasets/{dataset_id}/preview").json()
    assert preview["total_rows"] == 4
    assert len(preview["rows"]) == 4

    stats = client.get(f"/datasets/{dataset_id}/stats").json()
    assert stats["rows"] == 4
    assert stats["duplicate_rows"] == 1
    assert stats["missing_cells"] == 3

    insights = client.get(f"/datasets/{dataset_id}/insights").json()
    assert any(i["code"] == "duplicate_rows" for i in insights["issues"])


def test_upload_rejects_unsupported_extension(client: TestClient) -> None:
    res = _upload(client, name="notes.txt")
    assert res.status_code == 400
    assert "Unsupported file type" in res.json()["detail"]


def test_upload_rejects_unparseable_file(client: TestClient) -> None:
    res = _upload(client, content=b"definitely-not-a-zip", name="broken.xlsx")
    assert res.status_code == 400
    assert "Could not parse" in res.json()["detail"]


def test_upload_rejects_oversized_file(client: TestClient) -> None:
    settings = get_settings()
    original = settings.max_upload_mb
    settings.max_upload_mb = 0
    try:
        res = _upload(client)
        assert res.status_code == 400
        assert "upload limit" in res.json()["detail"]
    finally:
        settings.max_upload_mb = original


def test_job_lifecycle_completes_and_exports(client: TestClient) -> None:
    dataset_id = _upload(client).json()["id"]

    created = client.post(
        "/cleaning/jobs",
        json={
            "dataset_id": dataset_id,
            "operations": [{"operation": "remove_duplicates", "params": {}}],
        },
    )
    assert created.status_code == 202
    job = created.json()
    assert job["status"] == "pending"

    # TestClient ran the background task already; the job must be terminal.
    polled = client.get(f"/cleaning/jobs/{job['id']}").json()
    assert polled["status"] == "completed"
    assert polled["error"] is None
    assert polled["finished_at"] is not None

    preview = client.get(f"/cleaning/jobs/{job['id']}/preview").json()
    assert preview["total_rows"] == 3  # duplicate row removed

    report = client.get(f"/reports/by-job/{job['id']}").json()
    assert report["payload"]["rows_before"] == 4
    assert report["payload"]["rows_after"] == 3

    download = client.get(f"/exports/{job['id']}/download?format=csv")
    assert download.status_code == 200
    assert "attachment" in download.headers["content-disposition"]
    assert download.text.count("\n") >= 3

    excel = client.get(f"/exports/{job['id']}/download?format=excel")
    assert excel.status_code == 200
    assert excel.content[:2] == b"PK"  # xlsx is a zip container


def test_job_with_unknown_operation_is_rejected(client: TestClient) -> None:
    dataset_id = _upload(client).json()["id"]
    res = client.post(
        "/cleaning/jobs",
        json={
            "dataset_id": dataset_id,
            "operations": [{"operation": "no_such_op", "params": {}}],
        },
    )
    assert res.status_code == 400
    assert "Unknown operation" in res.json()["detail"]


def _add_job(dataset_id: int, status: JobStatus, created_at: datetime | None = None) -> int:
    session = SessionLocal()
    try:
        job = CleaningJob(dataset_id=dataset_id, status=status, operations=[])
        if created_at is not None:
            job.created_at = created_at
        session.add(job)
        session.commit()
        return job.id
    finally:
        session.close()


def _job_status(job_id: int) -> tuple[str, str | None]:
    session = SessionLocal()
    try:
        job = session.get(CleaningJob, job_id)
        return job.status.value, job.error
    finally:
        session.close()


def _report_count(job_id: int) -> int:
    session = SessionLocal()
    try:
        return session.scalar(
            select(func.count()).select_from(Report).where(Report.job_id == job_id)
        )
    finally:
        session.close()


def test_recover_stale_jobs_fails_orphans(client: TestClient) -> None:
    dataset_id = _upload(client).json()["id"]
    old = datetime(2020, 1, 1)
    stale_pending = _add_job(dataset_id, JobStatus.pending, created_at=old)
    stale_running = _add_job(dataset_id, JobStatus.running, created_at=old)
    fresh_pending = _add_job(dataset_id, JobStatus.pending)

    session = SessionLocal()
    try:
        recovered = CleaningService(session).recover_stale_jobs(
            max_age=timedelta(minutes=30)
        )
    finally:
        session.close()

    assert recovered >= 2
    for job_id in (stale_pending, stale_running):
        status, error = _job_status(job_id)
        assert status == "failed"
        assert "restart" in (error or "").lower()
    # A job younger than max_age is left for the worker to pick up.
    status, _ = _job_status(fresh_pending)
    assert status == "pending"


def test_run_skips_non_pending_jobs_and_never_duplicates_reports(
    client: TestClient,
) -> None:
    dataset_id = _upload(client).json()["id"]
    created = client.post(
        "/cleaning/jobs",
        json={
            "dataset_id": dataset_id,
            "operations": [{"operation": "remove_duplicates", "params": {}}],
        },
    ).json()
    job_id = created["id"]
    assert _job_status(job_id)[0] == "completed"
    assert _report_count(job_id) == 1

    # A duplicate run call (e.g. a second worker) must not re-execute the job.
    session = SessionLocal()
    try:
        job = CleaningService(session).run(job_id)
        assert job.status == JobStatus.completed
    finally:
        session.close()
    assert _report_count(job_id) == 1

    # A job already claimed by another worker is skipped as well.
    contested = _add_job(dataset_id, JobStatus.pending)
    claimer = SessionLocal()
    try:
        assert JobRepository(claimer).claim_pending(contested) is True
        claimer.commit()
    finally:
        claimer.close()
    session = SessionLocal()
    try:
        job = CleaningService(session).run(contested)
        assert job.status == JobStatus.running  # untouched by the loser
    finally:
        session.close()
    assert _report_count(contested) == 0


def test_sidecar_write_failure_does_not_fail_job(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    dataset_id = _upload(client).json()["id"]

    def boom(*_args, **_kwargs):
        raise OSError("disk full")

    monkeypatch.setattr("app.io.metadata.write_meta", boom)
    created = client.post(
        "/cleaning/jobs",
        json={
            "dataset_id": dataset_id,
            "operations": [{"operation": "remove_duplicates", "params": {}}],
        },
    )
    assert created.status_code == 202
    job_id = created.json()["id"]

    status, error = _job_status(job_id)
    assert status == "completed"
    assert error is None
    # Preview falls back to reading the cleaned CSV when the cache is absent.
    preview = client.get(f"/cleaning/jobs/{job_id}/preview")
    assert preview.status_code == 200
    assert preview.json()["total_rows"] == 3


def test_cached_endpoints_survive_reader_failure(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Proves preview/stats/insights come from the sidecar, not a re-read."""
    dataset_id = _upload(client).json()["id"]

    def boom(*_args, **_kwargs):
        raise AssertionError("read_stored must not be called for cached data")

    monkeypatch.setattr("app.io.readers.read_stored", boom)

    preview = client.get(f"/datasets/{dataset_id}/preview")
    assert preview.status_code == 200
    assert preview.json()["total_rows"] == 4

    stats = client.get(f"/datasets/{dataset_id}/stats")
    assert stats.status_code == 200
    assert stats.json()["duplicate_rows"] == 1

    insights = client.get(f"/datasets/{dataset_id}/insights")
    assert insights.status_code == 200
    assert any(i["code"] == "duplicate_rows" for i in insights.json()["issues"])


FORMULA_CSV = (
    b"name,note\n"
    b"Alice,=1+1\n"
    b"Bob,+2\n"
    b"Carol,-3\n"
    b"Dan,@SUM(A1:A2)\n"
    b"Eve,ok\n"
)


def test_escape_formula_injection_unit() -> None:
    df = pd.DataFrame(
        {
            "text": ["=1+1", "+2", "-3", "@x", "\tx", "\rx", "ok"],
            "number": [-5, -5, -5, -5, -5, -5, -5],
        }
    )
    out = writers.escape_formula_injection(df)
    assert out["text"].tolist() == [
        "'=1+1",
        "'+2",
        "'-3",
        "'@x",
        "'\tx",
        "'\rx",
        "ok",
    ]
    # Numeric values are never touched.
    assert out["number"].tolist() == [-5, -5, -5, -5, -5, -5, -5]


def test_csv_and_xlsx_download_escape_formulas(client: TestClient) -> None:
    dataset_id = client.post(
        "/datasets/upload",
        files={"file": ("danger.csv", FORMULA_CSV, "text/csv")},
    ).json()["id"]
    job = client.post(
        "/cleaning/jobs", json={"dataset_id": dataset_id, "operations": []}
    ).json()
    assert client.get(f"/cleaning/jobs/{job['id']}").json()["status"] == "completed"

    csv_text = client.get(f"/exports/{job['id']}/download?format=csv").text
    for payload in ("'=1+1", "'+2", "'-3", "'@SUM(A1:A2)"):
        assert payload in csv_text
    assert "ok" in csv_text and "'ok" not in csv_text

    xlsx = client.get(f"/exports/{job['id']}/download?format=excel")
    cells = [
        c.value
        for row in load_workbook(io.BytesIO(xlsx.content)).active.iter_rows()
        for c in row
    ]
    assert "'=1+1" in cells and "'@SUM(A1:A2)" in cells


def test_sql_import_disabled_by_default(client: TestClient) -> None:
    res = client.post(
        "/datasets/import-sql",
        json={
            "name": "x",
            "connection_url": "sqlite://",
            "query": "SELECT 1",
        },
    )
    assert res.status_code == 403
    assert "disabled" in res.json()["detail"].lower()


def test_sql_import_round_trip_when_enabled(client: TestClient) -> None:
    db_path = Path(tempfile.mkdtemp()) / "source.db"
    conn = sqlite3.connect(db_path)
    conn.execute("CREATE TABLE t (a INTEGER, b TEXT)")
    conn.executemany("INSERT INTO t VALUES (?, ?)", [(1, "x"), (2, "y")])
    conn.commit()
    conn.close()

    settings = get_settings()
    settings.enable_sql_import = True
    try:
        res = client.post(
            "/datasets/import-sql",
            json={
                "name": "from_sql",
                "connection_url": f"sqlite:///{db_path.as_posix()}",
                "query": "SELECT a, b FROM t",
            },
        )
    finally:
        settings.enable_sql_import = False

    assert res.status_code == 201
    body = res.json()
    assert body["row_count"] == 2
    assert body["column_count"] == 2


def test_request_id_is_generated_and_echoed(client: TestClient) -> None:
    # Generated when absent.
    res = client.get("/health")
    generated = res.headers.get("X-Request-ID")
    assert generated and len(generated) >= 16

    # Reused when supplied by the caller.
    res = client.get("/health", headers={"X-Request-ID": "trace-abc-123"})
    assert res.headers["X-Request-ID"] == "trace-abc-123"


def test_download_requires_completed_job(client: TestClient) -> None:
    dataset_id = _upload(client).json()["id"]
    session = SessionLocal()
    try:
        job = CleaningJob(
            dataset_id=dataset_id, status=JobStatus.pending, operations=[]
        )
        session.add(job)
        session.commit()
        job_id = job.id
    finally:
        session.close()

    res = client.get(f"/exports/{job_id}/download?format=csv")
    assert res.status_code == 400
    assert "not completed" in res.json()["detail"]

    preview = client.get(f"/cleaning/jobs/{job_id}/preview")
    assert preview.status_code == 400
